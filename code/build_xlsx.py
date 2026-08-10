import pandas as pd, numpy as np, json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

R=json.load(open("results.json"))
o=pd.read_csv("data/orders.csv"); b=pd.read_csv("data/buyers.csv")
s=pd.read_csv("data/shows.csv"); rfm=pd.read_csv("data/rfm_segments.csv")
orb=o[o.tour=="ORBITE 2026-27"]

INK="FF141414"; ACC="FFC4452B"; LIGHT="FFF2F2F2"
H=Font(name="Arial",size=10,bold=True,color="FFFFFFFF")
HF=PatternFill("solid",fgColor=INK)
T=Font(name="Arial",size=14,bold=True,color=INK)
N=Font(name="Arial",size=10); NB=Font(name="Arial",size=10,bold=True)
INP=Font(name="Arial",size=10,color="FF0000FF")
MUT=Font(name="Arial",size=9,color="FF7A7A7A",italic=True)
thin=Border(bottom=Side("thin",color="FFDDDDDD"))

wb=Workbook(); wb.remove(wb.active)

def head(ws,row,cols,widths):
    for i,(c,w) in enumerate(zip(cols,widths),1):
        cell=ws.cell(row,i,c); cell.font=H; cell.fill=HF
        cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width=w
    ws.row_dimensions[row].height=28

# ---------------------------------------------------------------- LISEZ-MOI
ws=wb.create_sheet("Lisez-moi")
ws.column_dimensions["A"].width=104
rows=[("LÜMA — Tournée ORBITE 2026-27",T),
 ("Suivi de billetterie et segmentation CRM · analyse arrêtée au 4 août 2026",N),(None,None),
 ("Ce classeur est un exercice de portfolio. Les données sont simulées par un script Python à graine fixe",MUT),
 ("(generate.py) : l'artiste, la tournée et les acheteurs sont fictifs. La méthode, elle, ne l'est pas.",MUT),(None,None),
 ("Onglets",NB),
 ("Rythme de vente     remplissage par date, rapporté à la courbe de référence de la tournée",N),
 ("Canaux              mix de distribution et part d'audience réellement joignable",N),
 ("Segments            segmentation RFM des acheteurs, effectifs et valeur",N),
 ("Plan de campagne    cible, test A/B et prévision par campagne — cellules bleues = hypothèses modifiables",N),(None,None),
 ("Convention de lecture",NB),
 ("Bleu = hypothèse que vous pouvez changer. Noir = calculé par formule.",N),
 ("Indice de rythme < 0,85 = date en retard sur la médiane de la tournée au même stade de commercialisation.",N)]
for i,(t,f) in enumerate(rows,1):
    if t: c=ws.cell(i,1,t); c.font=f

# ------------------------------------------------------------ RYTHME DE VENTE
ws=wb.create_sheet("Rythme de vente")
ws["A1"]="Rythme de vente par date"; ws["A1"].font=T
ws["A2"]="Indice = remplissage réel ÷ remplissage médian de la tournée au même nombre de jours en vente."
ws["A2"].font=MUT
head(ws,4,["Ville","Salle","Jauge","Billets vendus","Remplissage","Indice de rythme","Jours avant concert","Statut"],
     [15,24,10,15,13,15,15,16])
pace=pd.DataFrame(R["pace"])
for i,r in enumerate(pace.itertuples(),5):
    ws.cell(i,1,r.city).font=N; ws.cell(i,2,r.venue).font=N
    ws.cell(i,3,int(r.capacity)).font=N; ws.cell(i,4,int(r.sold)).font=N
    ws.cell(i,5,f"=D{i}/C{i}").font=N; ws.cell(i,5).number_format="0.0%"
    ws.cell(i,6,round(float(r.pace),2)).font=N; ws.cell(i,6).number_format="0.00"
    ws.cell(i,7,int(r.days_to_show)).font=N
    ws.cell(i,8,f'=IF(F{i}<$B$19,"En retard",IF(F{i}>1.1,"En avance","Conforme"))').font=N
    for col in range(1,9): ws.cell(i,col).border=thin
last=4+len(pace)
ws.cell(last+1,3,f"=SUM(C5:C{last})").font=NB; ws.cell(last+1,4,f"=SUM(D5:D{last})").font=NB
ws.cell(last+1,1,"Total tournée").font=NB
ws.cell(last+1,5,f"=D{last+1}/C{last+1}").font=NB; ws.cell(last+1,5).number_format="0.0%"
ws.cell(last+3,1,"Seuil d'alerte sur l'indice").font=N
ws.cell(last+3,2,0.85).font=INP; ws.cell(last+3,2).number_format="0.00"
ws.cell(last+4,1,"Billets manquants sur les dates en retard, pour revenir à la médiane").font=N
ws.cell(last+4,2,R["gap"]).font=INP
ws.cell(last+6,1,"Hypothèse : la médiane de la tournée sert de référence faute d'historique 2024 date par date.").font=MUT

# ------------------------------------------------------------------- CANAUX
ws=wb.create_sheet("Canaux")
ws["A1"]="Mix de distribution et captation de la donnée"; ws["A1"].font=T
ws["A2"]="Les revendeurs restituent le chiffre d'affaires, pas l'adresse e-mail du fan."; ws["A2"].font=MUT
head(ws,4,["Canal","Billets","Part","CA billetterie (€)","Frais payés par le fan (€)","Adresse e-mail restituée"],
     [20,12,10,20,24,24])
ch=pd.DataFrame(R["channels"])
for i,r in enumerate(ch.itertuples(),5):
    ws.cell(i,1,r.channel).font=N; ws.cell(i,2,int(r.tickets)).font=N
    ws.cell(i,3,f"=B{i}/$B${5+len(ch)}").font=N; ws.cell(i,3).number_format="0.0%"
    ws.cell(i,4,round(float(r.gross))).font=N; ws.cell(i,4).number_format="#,##0"
    ws.cell(i,5,round(float(r.fees))).font=N;  ws.cell(i,5).number_format="#,##0"
    ws.cell(i,6,"Oui" if r.channel=="Site officiel" else "Non").font=N
    for c in range(1,7): ws.cell(i,c).border=thin
e=4+len(ch)
for col,lab in ((1,"Total"),):
    ws.cell(e+1,col,lab).font=NB
for col in (2,4,5): ws.cell(e+1,col,f"=SUM({get_column_letter(col)}5:{get_column_letter(col)}{e})").font=NB
ws.cell(e+1,4).number_format="#,##0"; ws.cell(e+1,5).number_format="#,##0"
ws.cell(e+3,1,"Acheteurs uniques de la tournée").font=N; ws.cell(e+3,2,R["orbite_buyers"]).font=N
ws.cell(e+4,1,"Dont joignables (e-mail détenu en propre + opt-in)").font=N; ws.cell(e+4,2,R["contactable"]).font=N
ws.cell(e+5,1,"Part joignable").font=NB; ws.cell(e+5,2,f"=B{e+4}/B{e+3}").font=NB
ws.cell(e+5,2).number_format="0.0%"
ws.cell(e+6,1,"Acheteurs hors de portée").font=N; ws.cell(e+6,2,f"=B{e+3}-B{e+4}").font=N

# ----------------------------------------------------------------- SEGMENTS
ws=wb.create_sheet("Segments")
ws["A1"]="Segmentation RFM des acheteurs"; ws["A1"].font=T
ws["A2"]="Récence, fréquence et montant calculés sur l'historique complet, tournée 2024 incluse."; ws["A2"].font=MUT
head(ws,4,["Segment","Acheteurs","Récence médiane (j)","Commandes / acheteur","Dépense moyenne (€)",
           "Part joignable","Cible prioritaire"],[24,12,20,20,20,14,18])
sg=pd.DataFrame(R["segments"])
prio={"À réactiver":"Oui — campagne C","Noyau dur":"Oui — prévente","Multi-dates":"Oui — parrainage",
      "Gros paniers":"Non — déjà convertis","Nouveaux acheteurs":"Non — nurturing"}
for i,r in enumerate(sg.itertuples(),5):
    ws.cell(i,1,r.segment).font=N; ws.cell(i,2,int(r.buyers)).font=N
    ws.cell(i,3,round(float(r.recency))).font=N
    ws.cell(i,4,round(float(r.freq),2)).font=N; ws.cell(i,4).number_format="0.00"
    ws.cell(i,5,round(float(r.spend),1)).font=N; ws.cell(i,5).number_format="#,##0.0"
    ws.cell(i,6,round(float(r.contactable),3)).font=N; ws.cell(i,6).number_format="0.0%"
    ws.cell(i,7,prio.get(r.segment,"")).font=N
    for c in range(1,8): ws.cell(i,c).border=thin
q=4+len(sg)
ws.cell(q+1,1,"Total").font=NB; ws.cell(q+1,2,f"=SUM(B5:B{q})").font=NB
ws.cell(q+3,1,"Règles de segmentation").font=NB
for j,txt in enumerate([
  "Noyau dur : présent sur la tournée 2024 ET sur ORBITE.",
  "Multi-dates : au moins deux villes différentes sur ORBITE.",
  "Gros paniers : dépense dans le quartile supérieur, une seule tournée.",
  "Nouveaux acheteurs : première commande sur ORBITE.",
  "À réactiver : venu en 2024, aucune commande depuis."],1):
    ws.cell(q+3+j,1,txt).font=N

# ----------------------------------------------------------- PLAN DE CAMPAGNE
ws=wb.create_sheet("Plan de campagne")
ws["A1"]="Plan de campagne — prévision"; ws["A1"].font=T
ws["A2"]="Cellules bleues = hypothèses. Changez-les, la prévision se recalcule."; ws["A2"].font=MUT
head(ws,4,["Campagne","Cible adressable","Taux de clic","Clic → achat","Acheteurs attendus",
           "Billets / commande","Billets attendus","Panier moyen (€)","CA attendu (€)"],
     [30,16,12,13,16,16,15,15,15])
camp=[("A · Rattrapage géolocalisé (3 dates)",462,0.22,0.09,1.87,56.15),
      ("B · Réactivation Premier Cercle",339,0.28,0.11,1.87,56.15),
      ("C · Prévente Noyau dur (dates en retard)",118,0.41,0.18,1.87,56.15)]
for i,(n,t,ctr,cvr,tpo,bask) in enumerate(camp,5):
    ws.cell(i,1,n).font=N
    ws.cell(i,2,t).font=INP
    ws.cell(i,3,ctr).font=INP; ws.cell(i,3).number_format="0.0%"
    ws.cell(i,4,cvr).font=INP; ws.cell(i,4).number_format="0.0%"
    ws.cell(i,5,f"=B{i}*C{i}*D{i}").font=N; ws.cell(i,5).number_format="#,##0"
    ws.cell(i,6,tpo).font=INP; ws.cell(i,6).number_format="0.00"
    ws.cell(i,7,f"=E{i}*F{i}").font=N; ws.cell(i,7).number_format="#,##0"
    ws.cell(i,8,bask).font=INP; ws.cell(i,8).number_format="#,##0.00"
    ws.cell(i,9,f"=E{i}*H{i}").font=N; ws.cell(i,9).number_format="#,##0"
    for c in range(1,10): ws.cell(i,c).border=thin
t=4+len(camp)
ws.cell(t+1,1,"Total").font=NB
for c in (5,7,9): ws.cell(t+1,c,f"=SUM({get_column_letter(c)}5:{get_column_letter(c)}{t})").font=NB
ws.cell(t+1,7).number_format="#,##0"; ws.cell(t+1,9).number_format="#,##0"
ws.cell(t+3,1,"Écart à combler sur les dates en retard (billets)").font=N
ws.cell(t+3,2,R["gap"]).font=INP
ws.cell(t+4,1,"Part de l'écart couverte par le CRM").font=NB
ws.cell(t+4,2,f"=G{t+1}/B{t+3}").font=NB; ws.cell(t+4,2).number_format="0.0%"
for j,txt in enumerate([
 "",
 "Lecture : le CRM seul ne peut pas combler l'écart. La base joignable est trop petite —",
 "c'est le constat central de l'étude, pas un échec de la campagne. Le levier structurel est",
 "la captation d'e-mails au moment de la vente, traitée en campagne D (voir la page du projet).",
 "",
 "Hypothèses de taux : ordres de grandeur de marché pour une audience chaude et ciblée.",
 "À remplacer par les taux réels de l'outil d'envoi dès la première campagne exécutée.",
 "Chaque campagne prévoit un groupe témoin de 10 % non sollicité : sans témoin, on mesure",
 "la saisonnalité, pas l'effet de la campagne."],1):
    ws.cell(t+5+j,1,txt).font=MUT

wb.save("/tmp/pf/out/LUMA_ORBITE_suivi_billetterie.xlsx")
print("saved")
